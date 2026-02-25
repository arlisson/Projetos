"""
Aplicativo desktop para cadastro de leads em planilha Excel (XLSX).

Funcionalidades:
- Campos fixos: Nome, Telefone, Email
- Campos dinâmicos: criar, editar título e tipo e excluir (reflete em colunas do Excel)
- Tipos fechados (lista): texto, telefone, email, numero, moeda, data, booleano
- Planilha persistida: caminho salvo em controle_planilha.json
- UI customizável: ícones e tema via controle_ui.json
- Tema com sliders: usuário ajusta a cor de fundo (H/S/L); demais cores derivadas automaticamente
- Ícones tintados: ícones originalmente pretos passam a seguir a cor do tema
- Salvamento: adiciona linha nova (append) sem sobrescrever
- Ao salvar: limpa automaticamente os campos
- NOVO: ao selecionar uma planilha já existente, opção de substituir os campos do app pelos cabeçalhos (linha 1) da planilha

Arquivos de controle:
- controle_campos.json
- controle_planilha.json
- controle_ui.json

Dependências:
pip install PySide6 openpyxl
"""

from user_login import prompt_email, clear_login
from online_license import ensure_online_license, clear_cache
from simple_lock import ensure_or_mark
from loading_screen import LoadingScreen

# considere: device_limit, no_license, blocked, not_activated, revoked
EMAIL_RETRY_CODES = {"no_license", "blocked", "device_limit", "not_activated", "revoked"}

import json
import sys
import os
import colorsys
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QIcon, QPixmap, QColor, QPainter
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QVBoxLayout,
    QWidget,
    QDialog,
    QSlider,
    QFormLayout,
    QDialogButtonBox,
    QFrame
)

CONFIG_FIELDS_PATH = "controle_campos.json"
CONFIG_SHEET_PATH = "controle_planilha.json"
CONFIG_UI_PATH = "controle_ui.json"

# =========================
# Tipos fechados (metadados)
# =========================

FIELD_TYPES: List[str] = [
    "texto",
    "telefone",
    "email",
    "numero",
    "moeda",
    "data",
    "booleano",
]

EXCEL_NUMBER_FORMAT: Dict[str, str] = {
    "texto": "@",
    "telefone": "@",
    "email": "@",
    "numero": "0.00",
    "moeda": '"R$" #,##0.00',
    "data": "dd/mm/yyyy",
    "booleano": "@",
}


# =========================
# Utilitários de caminho
# =========================

def abs_path(p: str) -> str:
    if not p:
        return ""
    if os.path.isabs(p):
        return p
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, p)


def file_exists(p: str) -> bool:
    ap = abs_path(p)
    return bool(ap) and os.path.exists(ap)


# =========================
# Cores / tema
# =========================

DEFAULT_THEME = {
    "background": "#0B1220",
    "surface": "#0F1A2B",
    "surface_alt": "#111F33",
    "text": "#E6EDF7",
    "muted_text": "#A7B3C6",
    "primary": "#3B82F6",
    "danger": "#EF4444",
    "border": "#1F2A44",
}


def hex_to_rgb(hex_color: str) -> Tuple[int, int, int]:
    h = (hex_color or "").lstrip("#")
    if len(h) != 6:
        return (0, 0, 0)
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def rgb_to_hex(r: int, g: int, b: int) -> str:
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"#{r:02X}{g:02X}{b:02X}"


def rgb_to_hsl(r: int, g: int, b: int) -> Tuple[int, int, int]:
    rf, gf, bf = r / 255.0, g / 255.0, b / 255.0
    h, l, s = colorsys.rgb_to_hls(rf, gf, bf)
    return int(round(h * 359)), int(round(s * 100)), int(round(l * 100))


def hsl_to_rgb(h: int, s: int, l: int) -> Tuple[int, int, int]:
    hf = (h % 360) / 360.0
    sf = max(0, min(100, s)) / 100.0
    lf = max(0, min(100, l)) / 100.0
    r, g, b = colorsys.hls_to_rgb(hf, lf, sf)
    return int(round(r * 255)), int(round(g * 255)), int(round(b * 255))


def luminance(rgb: Tuple[int, int, int]) -> float:
    r, g, b = rgb
    return (0.2126 * r + 0.7152 * g + 0.0722 * b) / 255.0


def blend(a: Tuple[int, int, int], b: Tuple[int, int, int], t: float) -> Tuple[int, int, int]:
    t = max(0.0, min(1.0, t))
    return (
        int(round(a[0] * (1 - t) + b[0] * t)),
        int(round(a[1] * (1 - t) + b[1] * t)),
        int(round(a[2] * (1 - t) + b[2] * t)),
    )

def get_sheet_headers(xlsx_path: str, sheet_name: str) -> list[str]:
    """
    Extract the header row from an Excel worksheet.
    
    This function reads the first row of a specified worksheet in an Excel file
    and returns a list of non-empty header values as strings.
    
    Args:
        xlsx_path (str): The file path to the Excel workbook (.xlsx file).
        sheet_name (str): The name of the worksheet to read headers from.
            If the sheet name does not exist in the workbook, the active sheet is used.
    
    Returns:
        list[str]: A list of header values from the first row, with whitespace stripped.
            Only non-empty values are included in the returned list.
    
    Example:
        >>> headers = get_sheet_headers('data.xlsx', 'Sheet1')
        >>> print(headers)
        ['Name', 'Email', 'Phone']
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = []
    for cell in ws[1]:
        v = (cell.value or "")
        v = str(v).strip()
        if v:
            headers.append(v)
    return headers

def normalize_masked_text(value: str) -> str:
    """
    Normalize and validate a masked text input by removing values that contain only separators or incomplete patterns.
    This function processes text that may contain data entry masks (like date or phone number formats)
    and returns an empty string if the input appears to be incomplete or contains only placeholder characters.
    Args:
        value (str): The input text to normalize, potentially containing masks or separators.
    Returns:
        str: The normalized input string if it contains meaningful data, or an empty string if it
             consists only of separators, placeholders, or incomplete masked patterns.
    Logic:
        - Returns empty string if input contains only separators/placeholders (e.g., "//", "__/__/____").
        - Returns empty string if input contains "/" but no digits (likely an incomplete date/mask).
        - Returns empty string if input contains "/" but has fewer than 8 digits (incomplete date format like "12/__/____").
        - Otherwise, returns the normalized (stripped) input value.
    Examples:
        >>> normalize_masked_text("  /  /  ")
        ''
        >>> normalize_masked_text("12/__/____")
        ''
        >>> normalize_masked_text("12/10/2023")
        '12/10/2023'
    """
    v = (value or "").strip()

    # Se ficar só com separadores/placeholder (ex.: "//", "__/__/____", "  /  /    "), considere vazio
    only_separators = re.sub(r"[0-9A-Za-zÀ-ÿ]", "", v)  # remove letras e números
    if v and only_separators and all(ch in " _-./()[]" for ch in only_separators):
        # se não sobrou nenhum dígito/letra, é só máscara
        if not re.search(r"[0-9A-Za-zÀ-ÿ]", v):
            return ""

    # Caso específico comum: contém "/" mas não tem dígito nenhum
    if "/" in v and not re.search(r"\d", v):
        return ""

    # Se quiser ser mais estrito para data: se tiver menos de 8 dígitos, considere vazio
    # (ddmmaaaa = 8). Isso evita salvar "12/__/____"
    digits = re.sub(r"\D", "", v)
    if "/" in v and digits and len(digits) < 8:
        return ""

    return v

def derive_theme_from_background(bg_hex: str, base_theme: dict) -> dict:
    bg_rgb = hex_to_rgb(bg_hex)
    is_dark = luminance(bg_rgb) < 0.5

    white = (255, 255, 255)
    black = (0, 0, 0)

    if is_dark:
        surface_rgb = blend(bg_rgb, white, 0.06)
        surface_alt_rgb = blend(bg_rgb, white, 0.10)
        border_rgb = blend(bg_rgb, white, 0.16)
        text_rgb = hex_to_rgb(base_theme.get("text", "#E6EDF7"))
    else:
        surface_rgb = blend(bg_rgb, black, 0.06)
        surface_alt_rgb = blend(bg_rgb, black, 0.10)
        border_rgb = blend(bg_rgb, black, 0.16)
        text_rgb = (15, 23, 42)

    muted_rgb = blend(text_rgb, border_rgb, 0.55)

    out = dict(base_theme)
    out["background"] = bg_hex
    out["surface"] = rgb_to_hex(*surface_rgb)
    out["surface_alt"] = rgb_to_hex(*surface_alt_rgb)
    out["border"] = rgb_to_hex(*border_rgb)
    out["text"] = rgb_to_hex(*text_rgb)
    out["muted_text"] = rgb_to_hex(*muted_rgb)
    return out

def digits_only(value: str) -> str:
    # remove tudo que não é dígito (remove também _, espaços e símbolos da máscara)
    return re.sub(r"\D+", "", value or "")

def strip_mask_chars(value: str) -> str:
    # remove apenas placeholders/sobras comuns, preserva o resto
    # (útil se você quiser manter caracteres não numéricos em campos texto)
    return (value or "").replace("_", "").strip()

# =========================
# Modelos / Configuração
# =========================

@dataclass
class Campo:
    id: str
    titulo: str
    tipo: str = "texto"
    fixo: bool = False


def default_fields_config() -> dict:
    return {
        "arquivo_padrao": "PreencheFacil.xlsx",
        "aba": "Preenche Fácil",
        "campos": [
            {"id": "nome", "titulo": "Nome", "tipo": "texto", "fixo": True},
            {"id": "telefone", "titulo": "Telefone", "tipo": "telefone", "fixo": True},
            {"id": "email", "titulo": "Email", "tipo": "email", "fixo": True},
        ],
    }


def load_fields_config() -> dict:
    if not os.path.exists(CONFIG_FIELDS_PATH):
        cfg = default_fields_config()
        save_fields_config(cfg)
        return cfg

    with open(CONFIG_FIELDS_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    for c in cfg.get("campos", []):
        c.setdefault("tipo", "texto")
        c.setdefault("fixo", False)
        if c.get("tipo") not in FIELD_TYPES:
            c["tipo"] = "texto"

    return cfg


def save_fields_config(cfg: dict) -> None:
    with open(CONFIG_FIELDS_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def load_sheet_config() -> dict:
    if not os.path.exists(CONFIG_SHEET_PATH):
        return {}
    try:
        with open(CONFIG_SHEET_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_sheet_config(path: str) -> None:
    with open(CONFIG_SHEET_PATH, "w", encoding="utf-8") as f:
        json.dump({"last_sheet_path": path}, f, ensure_ascii=False, indent=2)


def get_last_sheet_path() -> Optional[str]:
    cfg = load_sheet_config()
    p = cfg.get("last_sheet_path")
    if not p:
        return None
    return p if os.path.exists(p) else None


def default_ui_config() -> dict:
    bg = DEFAULT_THEME["background"]
    r, g, b = hex_to_rgb(bg)
    h, s, l = rgb_to_hsl(r, g, b)
    return {
        "window_icon": "assets/A.ico",
        "button_icons": {
            "choose_file": "assets/icons/pasta.png",
            "add_field": "assets/icons/adicionar.png",
            "save_lead": "assets/icons/salvar.png",
            "clear": "assets/icons/limpar.png",
            "edit_title": "assets/icons/editar.png",
            "delete_field": "assets/icons/lixeira.png",
            "settings": "assets/icons/engrenagem.png",
        },
        "theme": dict(DEFAULT_THEME),
        "background_hsl": {"h": h, "s": s, "l": l},
    }


def load_ui_config() -> dict:
    if not os.path.exists(CONFIG_UI_PATH):
        cfg = default_ui_config()
        save_ui_config(cfg)
        return cfg
    try:
        with open(CONFIG_UI_PATH, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            cfg.setdefault("theme", dict(DEFAULT_THEME))
            cfg.setdefault("button_icons", {})
            if "background_hsl" not in cfg:
                r, g, b = hex_to_rgb(cfg["theme"].get("background", DEFAULT_THEME["background"]))
                h, s, l = rgb_to_hsl(r, g, b)
                cfg["background_hsl"] = {"h": h, "s": s, "l": l}
            return cfg
    except Exception:
        cfg = default_ui_config()
        save_ui_config(cfg)
        return cfg


def save_ui_config(cfg: dict) -> None:
    with open(CONFIG_UI_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# =========================
# Helpers: importação de campos do Excel
# =========================

def sanitize_id(title: str) -> str:
    s = (title or "").strip().lower()
    s = s.replace(" ", "_")
    s = re.sub(r"[^a-z0-9_]+", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "campo"


def make_unique_id(base_id: str, used: set) -> str:
    cid = base_id
    i = 2
    while cid in used:
        cid = f"{base_id}_{i}"
        i += 1
    used.add(cid)
    return cid

def is_excel_lock_present(xlsx_path: str) -> bool:
    folder = os.path.dirname(xlsx_path) or "."
    name = os.path.basename(xlsx_path)
    lock_name = "~$" + name
    return os.path.exists(os.path.join(folder, lock_name))


def infer_type_from_title(header: str) -> str:
    """
    Inferência simples (opcional) para ajudar.
    Você pode remover e retornar sempre "texto" se preferir.
    """
    h = (header or "").strip().lower()
    if "email" in h or "e-mail" in h:
        return "email"
    if "tel" in h or "fone" in h or "whats" in h or "cel" in h:
        return "telefone"
    if "data" in h or "dt" == h:
        return "data"
    if "preço" in h or "preco" in h or "valor" in h or "custo" in h:
        return "moeda"
    if "qtd" in h or "quant" in h or "numero" in h or "número" in h:
        return "numero"
    if "ativo" in h or "status" in h:
        return "booleano"
    return "texto"


def read_headers_from_excel(path: str, preferred_sheet: str) -> Tuple[str, List[str], bool]:
    """
    Lê a linha 1 e retorna (sheet_name_usado, headers, has_header).
    has_header = True se encontrou pelo menos 1 valor não vazio na linha 1.
    - Se preferred_sheet não existir, usa a aba ativa.
    """
    wb = load_workbook(path)
    sheet_name = preferred_sheet if preferred_sheet in wb.sheetnames else wb.active.title
    ws = wb[sheet_name]

    if ws.max_row < 1:
        return sheet_name, [], False

    headers: List[str] = []
    has_any = False

    for cell in ws[1]:
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            has_any = True
            if s not in headers:
                headers.append(s)

    return sheet_name, headers, has_any


# =========================
# Conversões por tipo (para salvar melhor no Excel)
# =========================

def parse_decimal_br(s: str) -> Optional[float]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    s = re.sub(r"[^\d,\.\-]+", "", s)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if "," in s:
            s = s.replace(".", "")
            s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def parse_date_any(s: str) -> Optional[date]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def normalize_bool_ptbr(s: str) -> Optional[str]:
    if s is None:
        return None
    s0 = str(s).strip().lower()
    if not s0:
        return None
    if s0 in {"sim", "s", "yes", "y", "true", "1"}:
        return "Sim"
    if s0 in {"não", "nao", "n", "no", "false", "0"}:
        return "Não"
    return str(s).strip()


def cast_value_by_type(tipo: str, raw: str) -> Any:
    raw = "" if raw is None else str(raw).strip()
    if raw == "":
        return ""
    if tipo == "numero":
        v = parse_decimal_br(raw)
        return v if v is not None else raw
    if tipo == "moeda":
        v = parse_decimal_br(raw)
        return v if v is not None else raw
    if tipo == "data":
        d = parse_date_any(raw)
        return d if d is not None else raw
    if tipo == "booleano":
        b = normalize_bool_ptbr(raw)
        return b if b is not None else ""
    return raw


# =========================
# Excel helpers (openpyxl)
# =========================

def ensure_workbook(path: str, sheet_name: str, headers: List[str]) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.max_row < 1:
        ws.append(headers)
        wb.save(path)
        return

    existing = [c.value for c in ws[1]]
    existing = [v for v in existing if v is not None]

    changed = False
    for h in headers:
        if h not in existing:
            existing.append(h)
            changed = True

    if changed:
        for col_idx, h in enumerate(existing, start=1):
            ws.cell(row=1, column=col_idx, value=h)

    wb.save(path)


def apply_column_type_rules(path: str, sheet_name: str, campos: List[Campo]) -> None:
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    headers = [h for h in headers if h is not None]
    header_to_col = {h: (headers.index(h) + 1) for h in headers}

    try:
        ws.data_validations.dataValidation = []
    except Exception:
        pass

    max_row = max(ws.max_row, 2)

    for c in campos:
        if c.titulo not in header_to_col:
            continue

        col_idx = header_to_col[c.titulo]
        col_letter = get_column_letter(col_idx)
        fmt = EXCEL_NUMBER_FORMAT.get(c.tipo, "@")

        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = fmt

        rng = f"{col_letter}2:{col_letter}1048576"

        if c.tipo == "numero":
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = "Digite um número válido (>= 0)."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "moeda":
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = "Digite um valor válido (>= 0)."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "data":
            dv = DataValidation(
                type="date",
                operator="between",
                formula1="DATE(1900,1,1)",
                formula2="DATE(2099,12,31)",
                allow_blank=True,
            )
            dv.errorTitle = "Data inválida"
            dv.error = "Digite uma data válida."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "booleano":
            dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = 'Use "Sim" ou "Não".'
            ws.add_data_validation(dv)
            dv.add(rng)

    wb.save(path)


def append_row_typed(path: str, sheet_name: str, campos: List[Campo], row_by_title: Dict[str, str]) -> None:
    headers = [c.titulo for c in campos]
    ensure_workbook(path, sheet_name, headers)

    wb = load_workbook(path)
    ws = wb[sheet_name]

    existing = [c.value for c in ws[1]]
    existing = [v for v in existing if v is not None]
    col_idx = {h: (existing.index(h) + 1) for h in existing}

    next_row = ws.max_row + 1

    for c in campos:
        h = c.titulo
        if h not in col_idx:
            continue

        value = cast_value_by_type(c.tipo, row_by_title.get(h, ""))
        cell = ws.cell(row=next_row, column=col_idx[h], value=value)
        cell.number_format = EXCEL_NUMBER_FORMAT.get(c.tipo, "@")

    wb.save(path)


def delete_column_by_header(path: str, sheet_name: str, header_name: str) -> None:
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    if header_name not in headers:
        wb.save(path)
        return

    col = headers.index(header_name) + 1
    ws.delete_cols(col, 1)
    wb.save(path)


def rename_column_header(path: str, sheet_name: str, old_header: str, new_header: str) -> None:
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    if old_header not in headers:
        wb.save(path)
        return

    col = headers.index(old_header) + 1
    ws.cell(row=1, column=col, value=new_header)
    wb.save(path)

def write_headers_from_campos(path: str, sheet_name: str, campos: List[Campo]) -> None:
    """
    Garante que a linha 1 tenha cabeçalhos (títulos) baseados em campos do app.
    Se a aba não existir, cria.
    """
    headers = [c.titulo for c in campos]

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # escreve cabeçalho a partir da coluna 1
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    wb.save(path)
# =========================
# Dialog: Configuração de fundo (sliders)
# =========================

class ThemeDialog(QDialog):
    def __init__(self, parent: QWidget, h: int, s: int, l: int):
        super().__init__(parent)
        self.setWindowTitle("Ajustar tema")
        self.setModal(True)
        self.setMinimumWidth(240)

        self.slider_h = QSlider(Qt.Horizontal)
        self.slider_s = QSlider(Qt.Horizontal)
        self.slider_l = QSlider(Qt.Horizontal)

        self.slider_h.setRange(0, 359)
        self.slider_s.setRange(0, 100)
        self.slider_l.setRange(0, 100)

        self.slider_h.setValue(h)
        self.slider_s.setValue(s)
        self.slider_l.setValue(l)

        self.lbl_preview = QLabel("Prévia")
        self.lbl_preview.setAlignment(Qt.AlignCenter)
        self.lbl_preview.setMinimumHeight(60)

        form = QFormLayout()
        form.addRow("Tom da cor", self.slider_h)
        form.addRow("Saturação", self.slider_s)
        form.addRow("Brilho", self.slider_l)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

        layout = QVBoxLayout()
        layout.addLayout(form)
        layout.addWidget(self.lbl_preview)
        layout.addWidget(btns)
        self.setLayout(layout)

        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)

    def values(self) -> Tuple[int, int, int]:
        return self.slider_h.value(), self.slider_s.value(), self.slider_l.value()

#Cursor sempre no início
class CursorStartLineEdit(QLineEdit):
    """
    QLineEdit que força o cursor para o início ao receber foco/clique.
    Útil para campos com inputMask (telefone, data) para facilitar colar.
    """
    def __init__(self, *args, force_cursor_start: bool = False, **kwargs):
        super().__init__(*args, **kwargs)
        self._force_cursor_start = force_cursor_start

    def set_force_cursor_start(self, enabled: bool) -> None:
        self._force_cursor_start = bool(enabled)

    def _move_cursor_to_start(self) -> None:
        if self._force_cursor_start:
            self.setCursorPosition(0)
            self.deselect()

    def focusInEvent(self, event):
        super().focusInEvent(event)
        if self._force_cursor_start:
            QTimer.singleShot(0, self._move_cursor_to_start)

    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        if self._force_cursor_start:
            QTimer.singleShot(0, self._move_cursor_to_start)

# =========================
# UI / Aplicação
# =========================

class App(QWidget):
    def __init__(self):
        super().__init__()

        self.cfg_fields = load_fields_config()
        self.cfg_ui = load_ui_config()

        self.sheet_name: str = self.cfg_fields.get("aba", "Preenche Fácil")
        self.file_path: Optional[str] = get_last_sheet_path()

        self.campos: List[Campo] = [Campo(**c) for c in self.cfg_fields.get("campos", [])]
        self.inputs: Dict[str, QLineEdit] = {}

        self._pixmap_cache: Dict[str, QPixmap] = {}

        self._apply_window_icon()
        self._build_ui()
        self._apply_theme_from_config()

        if self.file_path:
            # Mantém o comportamento padrão ao abrir (não importa automaticamente os campos)
            self._apply_file_path(self.file_path, prepare=True, silent=True)

    # ---------- ícones / tint ----------

    def _apply_window_icon(self) -> None:
        icon_path = self.cfg_ui.get("window_icon", "")
        if file_exists(icon_path):
            self.setWindowIcon(QIcon(abs_path(icon_path)))

    def _load_pixmap(self, path: str) -> Optional[QPixmap]:
        ap = abs_path(path)
        if not ap or not os.path.exists(ap):
            return None
        if ap in self._pixmap_cache:
            return self._pixmap_cache[ap]
        pm = QPixmap(ap)
        if pm.isNull():
            return None
        self._pixmap_cache[ap] = pm
        return pm

    def _tint_pixmap(self, pixmap: QPixmap, tint_hex: str) -> QPixmap:
        tinted = QPixmap(pixmap.size())
        tinted.fill(Qt.transparent)

        painter = QPainter(tinted)
        painter.drawPixmap(0, 0, pixmap)
        painter.setCompositionMode(QPainter.CompositionMode_SourceIn)
        painter.fillRect(tinted.rect(), QColor(tint_hex))
        painter.end()

        return tinted

    def _icon_for_key(self, key: str, tint_hex: str) -> Optional[QIcon]:
        icons = self.cfg_ui.get("button_icons", {}) or {}
        p = icons.get(key, "")
        if not p or not file_exists(p):
            return None
        pm = self._load_pixmap(p)
        if not pm:
            return None
        return QIcon(self._tint_pixmap(pm, tint_hex))

    def _apply_button_icon(self, btn: QPushButton, key: str, tint_hex: str) -> None:
        ic = self._icon_for_key(key, tint_hex)
        if ic:
            btn.setIcon(ic)

    # ---------- tema (cores) ----------

    def _apply_theme_from_config(self) -> None:
        base_theme = dict(DEFAULT_THEME)
        base_theme.update(self.cfg_ui.get("theme", {}) or {})

        hsl = self.cfg_ui.get("background_hsl", {}) or {}
        h = int(hsl.get("h", 210))
        s = int(hsl.get("s", 49))
        l = int(hsl.get("l", 8))

        r, g, b = hsl_to_rgb(h, s, l)
        bg_hex = rgb_to_hex(r, g, b)

        derived = derive_theme_from_background(bg_hex, base_theme)

        self.cfg_ui["theme"] = derived
        self.cfg_ui["background_hsl"] = {"h": h, "s": s, "l": l}
        save_ui_config(self.cfg_ui)

        self._apply_theme(derived)

    def _apply_theme(self, theme: dict) -> None:
        bg = theme["background"]
        surface = theme["surface"]
        surface_alt = theme["surface_alt"]
        text = theme["text"]
        muted = theme["muted_text"]
        primary = theme["primary"]
        danger = theme["danger"]
        border = theme["border"]

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {bg};
                color: {text};
                font-size: 13px;
            }}

            QLabel {{
                color: {text};
            }}

            QLineEdit {{
                background-color: {surface};
                color: {text};
                border: 1px solid {border};
                border-radius: 8px;
                padding: 8px;
            }}

            QScrollArea {{
                background-color: transparent;
                border: none;
            }}

            QScrollArea QWidget {{
                background-color: transparent;
            }}

            QPushButton {{
                background-color: {surface_alt};
                color: {text};
                border: 1px solid {border};
                border-radius: 10px;
                padding: 8px 10px;
            }}

            QPushButton:hover {{
                border-color: {primary};
            }}

            QPushButton:pressed {{
                background-color: {surface};
            }}

            QPushButton[variant="primary"] {{
                background-color: {primary};
                color: #FFFFFF;
                border: 1px solid {primary};
                font-weight: 600;
            }}

            QPushButton[variant="danger"] {{
                background-color: {danger};
                color: #FFFFFF;
                border: 1px solid {danger};
                font-weight: 600;
            }}

            QLabel#Status {{
                color: {muted};          

                
            }}

                    

        """)

        self._retint_all_icons(theme)

    def _retint_all_icons(self, theme: dict) -> None:
        text = theme["text"]
        white = "#FFFFFF"

        self._apply_button_icon(self.btn_file, "choose_file", text)
        self._apply_button_icon(self.btn_add_field, "add_field", text)
        self._apply_button_icon(self.btn_clear, "clear", text)
        self._apply_button_icon(self.btn_save, "save_lead", white)
        self._apply_button_icon(self.btn_settings, "settings", text)

        self._last_theme_for_fields = theme
        self.render_fields()

    # ---------- build UI ----------

    def _build_ui(self) -> None:
        self.setWindowTitle("Preenche Fácil - Avance")
        self.setMinimumWidth(490)

        root = QVBoxLayout()

        file_row = QHBoxLayout()

        self.btn_file = QPushButton("Escolher planilha…")
        self.lbl_file = QLabel("Arquivo: (não selecionado)")
        self.lbl_file.setWordWrap(True)
        self.btn_file.clicked.connect(self.choose_file)

        self.btn_settings = QPushButton("")
        self.btn_settings.setToolTip("Configurações de tema")
        self.btn_settings.setFixedWidth(44)
        self.btn_settings.clicked.connect(self.open_theme_settings)

        file_row.addWidget(self.btn_file)
        file_row.addWidget(self.lbl_file, 1)
        file_row.addWidget(self.btn_settings)
        root.addLayout(file_row)

        self.fields_container = QWidget()
        self.fields_layout = QVBoxLayout()
        self.fields_container.setLayout(self.fields_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.fields_container)
        root.addWidget(scroll, 1)

        actions = QHBoxLayout()

        self.btn_add_field = QPushButton("Adicionar Campo")
        self.btn_save = QPushButton("Salvar (nova linha)")
        self.btn_clear = QPushButton("Limpar")

        self.btn_save.setProperty("variant", "primary")

        actions.addWidget(self.btn_add_field)
        actions.addWidget(self.btn_clear)
        actions.addWidget(self.btn_save)

        self.btn_add_field.clicked.connect(self.add_field)
        self.btn_save.clicked.connect(self.save_lead)
        self.btn_clear.clicked.connect(self.clear_fields)

        root.addLayout(actions)

        self.lbl_status = QLabel("Preencha os campos (copie/cole) e clique em Salvar.")
        self.lbl_status.setWordWrap(True)
        self.lbl_status.setObjectName("Status")
        root.addWidget(self.lbl_status)

                # --- Rodapé corporativo ---
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Plain)
        sep.setFixedHeight(1)
        root.addWidget(sep)

        # --- Rodapé corporativo ---
                # --- Rodapé corporativo ---
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Plain)
        sep.setFixedHeight(1)
        root.addWidget(sep)

        footer_row = QHBoxLayout()

        self.lbl_footer_left = QLabel("<b>AVANCE Telefonia Empresarial<b> • WhatsApp (22) 98812-4656")
        # self.lbl_footer_right = QLabel("suporte@avance.com • (11) 99999-9999 • © 2026")

        self.lbl_footer_left.setObjectName("Footer")
        # self.lbl_footer_right.setObjectName("Footer")
        # self.lbl_footer_right.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        footer_row.addWidget(self.lbl_footer_left, 1)
        # footer_row.addWidget(self.lbl_footer_right, 1)

        footer_wrap = QWidget()
        footer_wrap.setLayout(footer_row)
        root.addWidget(footer_wrap)

        self.setLayout(root)

        self._last_theme_for_fields = dict(DEFAULT_THEME)
        self.render_fields()

        if self.file_path:
            self.lbl_file.setText(f"Arquivo: {self.file_path}")

    def _apply_input_mask(self, inp: QLineEdit, tipo: str) -> None:
        # habilita cursor no início apenas para data/telefone
        if isinstance(inp, CursorStartLineEdit):
            inp.set_force_cursor_start(tipo in ("telefone", "data"))

        if tipo == "telefone":
            inp.setInputMask("(00) 00000-0000;_")
        elif tipo == "data":
            inp.setInputMask("00/00/0000;_")
        else:
            inp.setInputMask("")

    def render_fields(self) -> None:
        # 1) Snapshot dos valores atuais antes de destruir os widgets
        previous_values: Dict[str, str] = {}
        for cid, widget in self.inputs.items():
            try:
                previous_values[cid] = widget.text()
            except Exception:
                previous_values[cid] = ""

        # 2) Limpa o layout e remove widgets
        while self.fields_layout.count():
            item = self.fields_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        # 3) Recria os campos
        self.inputs.clear()
        theme = getattr(self, "_last_theme_for_fields", dict(DEFAULT_THEME))
        text = theme.get("text", "#E6EDF7")
        white = "#FFFFFF"

        for campo in self.campos:
            row = QHBoxLayout()

            lbl = QLabel(f"{campo.titulo}  [{campo.tipo}]")
            lbl.setMinimumWidth(160)

            inp = CursorStartLineEdit()  # ou QLineEdit(), conforme seu código
            inp.setPlaceholderText(f"Digite ou cole: {campo.titulo}")

            # aplica máscara e comportamento do cursor (se você implementou)
            self._apply_input_mask(inp, campo.tipo)

            # 4) Restaura o valor anterior do mesmo campo.id (principal fix)
            inp.setText(previous_values.get(campo.id, ""))

            self.inputs[campo.id] = inp

            btn_edit = QPushButton("Editar")
            btn_del = QPushButton("Excluir")

            self._apply_button_icon(btn_edit, "edit_title", text)
            self._apply_button_icon(btn_del, "delete_field", white)

            btn_edit.clicked.connect(lambda _=False, cid=campo.id: self.edit_field(cid))
            btn_del.clicked.connect(lambda _=False, cid=campo.id: self.delete_field(cid))
            btn_del.setProperty("variant", "danger")

            row.addWidget(lbl)
            row.addWidget(inp, 1)
            row.addWidget(btn_edit)
            row.addWidget(btn_del)

            wrap = QWidget()
            wrap.setLayout(row)
            self.fields_layout.addWidget(wrap)

        self.fields_layout.addStretch(1)

    # ---------- persistência ----------

    def _persist_campos(self) -> None:
        self.cfg_fields["campos"] = [c.__dict__ for c in self.campos]
        self.cfg_fields["aba"] = self.sheet_name
        save_fields_config(self.cfg_fields)

    def _apply_file_path(self, path: str, prepare: bool, silent: bool = False) -> None:
        self.file_path = path
        self.lbl_file.setText(f"Arquivo: {path}")
        save_sheet_config(path)

        if prepare:
            try:
                headers = [c.titulo for c in self.campos]
                ensure_workbook(self.file_path, self.sheet_name, headers)
                apply_column_type_rules(self.file_path, self.sheet_name, self.campos)
                if not silent:
                    self.lbl_status.setText("Planilha preparada e salva como padrão.")
            except Exception as e:
                self.lbl_status.setText(f"Erro ao preparar planilha: {e}")

    # ---------- NOVO: sincronizar campos do app com cabeçalhos da planilha ----------

    def sync_fields_from_existing_excel(self, path: str) -> bool:
        """
        Lê a linha 1 da planilha e substitui os campos do app por esses cabeçalhos.
        Se NÃO houver cabeçalho, escreve os campos atuais (controle_campos.json) na linha 1
        e segue o fluxo sem importar.
        """
        try:
            sheet_used, headers, has_header = read_headers_from_excel(path, self.sheet_name)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao ler a planilha: {e}")
            return False

        # Se a aba configurada não existe, muda para a aba ativa utilizada.
        self.sheet_name = sheet_used
        self.cfg_fields["aba"] = self.sheet_name

        # NOVO: sem cabeçalho -> escreve os campos do controle na planilha e não substitui os campos do app
        if not has_header:
            try:
                write_headers_from_campos(path, self.sheet_name, self.campos)
                apply_column_type_rules(path, self.sheet_name, self.campos)
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao criar cabeçalho na planilha: {e}")
                return False

            self._persist_campos()
            self.render_fields()
            self.lbl_status.setText("Planilha sem cabeçalho: cabeçalhos foram criados a partir do controle_campos.json.")
            return True

        # fluxo atual: tem cabeçalho -> importa e substitui campos do app
        if not headers:
            QMessageBox.warning(
                self,
                "Sem cabeçalho",
                "A planilha não possui cabeçalho válido na linha 1."
            )
            return False

        existing_by_title = {c.titulo.strip().lower(): c for c in self.campos}

        used_ids: set = set()
        new_campos: List[Campo] = []

        for h in headers:
            key = h.strip().lower()
            old = existing_by_title.get(key)

            if old:
                cid = old.id
                tipo = old.tipo if old.tipo in FIELD_TYPES else "texto"
                fixo = old.fixo
                used_ids.add(cid)
            else:
                cid = make_unique_id(sanitize_id(h), used_ids)
                tipo = infer_type_from_title(h)
                if tipo not in FIELD_TYPES:
                    tipo = "texto"
                fixo = False

            new_campos.append(Campo(id=cid, titulo=h, tipo=tipo, fixo=fixo))

        self.campos = new_campos
        self._persist_campos()

        try:
            apply_column_type_rules(path, self.sheet_name, self.campos)
        except Exception:
            pass

        self.render_fields()
        self.lbl_status.setText("Campos importados da planilha e salvos no controle.")
        return True

    def choose_file(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Escolher planilha",
            self.cfg_fields.get("arquivo_padrao", "PreencheFacil.xlsx"),
            "Planilha Excel (*.xlsx)",
        )
        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        existed = os.path.exists(path)

        if existed:
            resp = QMessageBox.question(
                self,
                "Planilha existente",
                "Esta planilha já existe.\n\n"
                "Deseja importar os campos do cabeçalho (linha 1) para o aplicativo?\n\n"
                "Sim: substitui os campos do app pelos campos da planilha (linha 1).\n"
                "Não: mantém os campos atuais e garante as colunas no Excel.\n\n"
                "Obs.: se a planilha não tiver cabeçalho, o app criará os títulos automaticamente a partir do controle."
            )
            if resp == QMessageBox.Yes:
                # Não prepara (não altera cabeçalho). Apenas salva o caminho e importa campos.
                self._apply_file_path(path, prepare=False, silent=True)
                ok = self.sync_fields_from_existing_excel(path)
                if ok:
                    self.lbl_file.setText(f"Arquivo: {path}  (aba: {self.sheet_name})")
                return

        # Comportamento padrão: prepara planilha para conter os campos atuais
        self._apply_file_path(path, prepare=True, silent=False)

    # ---------- ações de campos ----------

    def add_field(self) -> None:
        title, ok = QInputDialog.getText(self, "Novo campo", "Título do campo:")
        if not ok or not title.strip():
            return

        title = title.strip()
        existing_titles = {c.titulo for c in self.campos}
        if title in existing_titles:
            QMessageBox.warning(self, "Atenção", "Já existe um campo com esse título.")
            return

        tipo, ok_tipo = QInputDialog.getItem(
            self,
            "Tipo do campo",
            "Selecione o tipo:",
            FIELD_TYPES,
            0,
            False
        )
        if not ok_tipo:
            return
        tipo = str(tipo)

        used_ids = {c.id for c in self.campos}
        cid = make_unique_id(sanitize_id(title), used_ids)

        self.campos.append(Campo(id=cid, titulo=title, tipo=tipo, fixo=False))
        self._persist_campos()

        self.render_fields()
        self.lbl_status.setText("Campo criado e salvo no arquivo de controle.")

        if self.file_path:
            try:
                headers = [c.titulo for c in self.campos]
                ensure_workbook(self.file_path, self.sheet_name, headers)
                apply_column_type_rules(self.file_path, self.sheet_name, self.campos)
            except Exception as e:
                self.lbl_status.setText(f"Campo criado, mas falhou ao atualizar Excel: {e}")

    def edit_field(self, field_id: str) -> None:
        campo = next((c for c in self.campos if c.id == field_id), None)
        if not campo:
            return

        # CAPTURA estado atual do campo antes de mudar qualquer coisa
        old_tipo = campo.tipo
        old_value = self.inputs[field_id].text() if field_id in self.inputs else ""

        new_title, ok = QInputDialog.getText(self, "Editar campo", "Título:", text=campo.titulo)
        if not ok:
            return
        new_title = new_title.strip()
        if not new_title:
            QMessageBox.warning(self, "Atenção", "Título não pode ser vazio.")
            return

        idx = FIELD_TYPES.index(campo.tipo) if campo.tipo in FIELD_TYPES else 0
        new_tipo, ok_tipo = QInputDialog.getItem(self, "Editar campo", "Tipo:", FIELD_TYPES, idx, False)
        if not ok_tipo:
            return
        new_tipo = str(new_tipo)

        existing_titles = {c.titulo for c in self.campos if c.id != field_id}
        if new_title in existing_titles:
            QMessageBox.warning(self, "Atenção", "Já existe um campo com esse título.")
            return

        old_title = campo.titulo
        campo.titulo = new_title
        campo.tipo = new_tipo
        self._persist_campos()

        try:
            if self.file_path:
                if old_title != new_title:
                    rename_column_header(self.file_path, self.sheet_name, old_title, new_title)
                apply_column_type_rules(self.file_path, self.sheet_name, self.campos)
        except Exception as e:
            self.lbl_status.setText(f"Atualizado no controle, mas falhou no Excel: {e}")

        # Recria UI (isso mantém os dados dos outros campos, conforme o snapshot do render_fields)
        self.render_fields()

        # LIMPA caracteres da máscara quando ela “sai”
        masked = {"telefone", "data"}
        inp = self.inputs.get(field_id)
        if inp:
            if old_tipo in masked and new_tipo not in masked:
                # máscara removida -> remove símbolos
                inp.setText(digits_only(old_value))
            elif old_tipo not in masked and new_tipo in masked:
                # máscara aplicada -> use apenas dígitos para preencher a máscara corretamente
                inp.setText(digits_only(old_value))
            else:
                # tipos sem máscara: opcionalmente remove "_" caso exista
                if new_tipo not in masked:
                    inp.setText(strip_mask_chars(inp.text()))

        self.lbl_status.setText("Campo atualizado (título/tipo).")

    def delete_field(self, field_id: str) -> None:
        campo = next((c for c in self.campos if c.id == field_id), None)
        if not campo:
            return

        msg = (
            f"Confirma excluir o campo '{campo.titulo}'?\n\n"
            "Isso também removerá a coluna correspondente na planilha Excel.\n"
            "Os dados dessa coluna serão removidos definitivamente."
        )
        if QMessageBox.question(self, "Confirmar exclusão", msg) != QMessageBox.Yes:
            return

        self.campos = [c for c in self.campos if c.id != field_id]
        self._persist_campos()

        try:
            if self.file_path:
                delete_column_by_header(self.file_path, self.sheet_name, campo.titulo)
                apply_column_type_rules(self.file_path, self.sheet_name, self.campos)
        except Exception as e:
            self.lbl_status.setText(f"Campo removido do controle, mas falhou ao remover coluna no Excel: {e}")

        self.render_fields()
        self.lbl_status.setText("Campo excluído (e coluna removida, quando aplicável).")

    # ---------- ações gerais ----------

    def clear_fields(self) -> None:
        for inp in self.inputs.values():
            inp.setText("")
        self.lbl_status.setText("Campos limpos.")

    def save_lead(self) -> None:
        if not self.file_path:
            QMessageBox.warning(self, "Atenção", "Selecione uma planilha primeiro.")
            return

        # 1) Aviso antecipado: planilha provavelmente aberta no Excel (arquivo de lock "~$")
        try:
            
            folder = os.path.dirname(self.file_path) or "."
            name = os.path.basename(self.file_path)
            if os.path.exists(os.path.join(folder, "~$" + name)):
                msg = (
                    "Não foi possível salvar porque a planilha parece estar aberta no Excel.\n\n"
                    "Feche o arquivo no Excel e tente novamente."
                )
                self.lbl_status.setText("Planilha aberta no Excel (bloqueada).")
                QMessageBox.warning(self, "Planilha em uso", msg)
                return
        except Exception:
            # Se a checagem falhar por algum motivo, não bloqueia o salvamento
            pass

        row_by_title: Dict[str, str] = {}
        for c in self.campos:
            raw = self.inputs[c.id].text().strip() if c.id in self.inputs else ""
            txt = normalize_masked_text(raw)
            row_by_title[c.titulo] = txt

        if all(not v for v in row_by_title.values()):
            QMessageBox.information(self, "Info", "Nada para salvar (todos os campos vazios).")
            return

        # 2) Tenta salvar com retry simples para varredura inicial de antivírus (lock temporário)
        import time
        last_err = None
        for attempt in range(6):  # ~0.2 + 0.4 + 0.8 + 1.6 + 3.2 + 6.4 = ~12.6s no pior caso
            try:
                append_row_typed(self.file_path, self.sheet_name, self.campos, row_by_title)
                last_err = None
                break
            except PermissionError as e:
                last_err = e
                # Se o Excel abriu/fechou e o lock apareceu, avisa com mensagem específica
                try:
                    import os
                    folder = os.path.dirname(self.file_path) or "."
                    name = os.path.basename(self.file_path)
                    if os.path.exists(os.path.join(folder, "~$" + name)):
                        msg = (
                            "Não foi possível salvar porque a planilha está aberta no Excel.\n\n"
                            "Feche o arquivo no Excel e tente novamente."
                        )
                        self.lbl_status.setText("Planilha aberta no Excel (bloqueada).")
                        QMessageBox.warning(self, "Planilha em uso", msg)
                        return
                except Exception:
                    pass

                # backoff
                time.sleep(0.2 * (2 ** attempt))
            except Exception as e:
                self.lbl_status.setText(f"Erro ao salvar: {e}")
                QMessageBox.critical(self, "Erro", f"Erro ao salvar: {e}")
                return

        if last_err is not None:
            msg = (
                "Não foi possível salvar a planilha.\n\n"
                "Possíveis causas:\n"
                "- O arquivo está aberto no Excel\n"
                "- O antivírus está verificando/bloqueando a gravação\n"
                "- A pasta/arquivo não permite escrita\n\n"
                "Feche o Excel (se estiver aberto), aguarde alguns segundos e tente novamente.\n"
                "Se o problema persistir, escolha outra pasta para a planilha."
            )
            self.lbl_status.setText(f"Erro ao salvar (arquivo bloqueado): {last_err}")
            QMessageBox.critical(self, "Erro ao salvar", msg)
            return

        self.clear_fields()
        self.lbl_status.setText("Dados salvos (nova linha adicionada). Campos limpos automaticamente.")
        QMessageBox.information(self, "Info", "Dados salvos (nova linha adicionada). Campos limpos automaticamente.")

    # ---------- configurações (engrenagem) ----------

    def open_theme_settings(self) -> None:
        hsl = self.cfg_ui.get("background_hsl", {}) or {}
        h = int(hsl.get("h", 210))
        s = int(hsl.get("s", 49))
        l = int(hsl.get("l", 8))

        before_cfg = json.loads(json.dumps(self.cfg_ui))
        before_theme = dict(self.cfg_ui.get("theme", dict(DEFAULT_THEME)))

        dlg = ThemeDialog(self, h, s, l)

        def on_change():
            hh, ss, ll = dlg.values()
            rr, gg, bb = hsl_to_rgb(hh, ss, ll)
            bg_hex = rgb_to_hex(rr, gg, bb)
            base = dict(DEFAULT_THEME)
            base.update(before_theme)
            derived = derive_theme_from_background(bg_hex, base)
            self._apply_theme(derived)

        dlg.slider_h.valueChanged.connect(on_change)
        dlg.slider_s.valueChanged.connect(on_change)
        dlg.slider_l.valueChanged.connect(on_change)

        on_change()

        if dlg.exec() == QDialog.Accepted:
            hh, ss, ll = dlg.values()
            rr, gg, bb = hsl_to_rgb(hh, ss, ll)
            bg_hex = rgb_to_hex(rr, gg, bb)

            base = dict(DEFAULT_THEME)
            base.update(before_theme)

            derived = derive_theme_from_background(bg_hex, base)
            self.cfg_ui = before_cfg
            self.cfg_ui["background_hsl"] = {"h": hh, "s": ss, "l": ll}
            self.cfg_ui["theme"] = derived
            save_ui_config(self.cfg_ui)

            self._apply_theme(derived)
        else:
            self.cfg_ui = before_cfg
            save_ui_config(self.cfg_ui)
            self._apply_theme(before_theme)


def main() -> None:
    app = QApplication([])

    loader = LoadingScreen(app, title="Validando acesso")
    loader.show("Preparando validação...")

    while True:
        loader.hide()
        email = prompt_email(force=False)
        if not email:
            loader.close()
            return

        loader.show("Validando licença online...")
        ok, msg, code = ensure_online_license(email)
        if ok:
            loader.update("Licença validada com sucesso.")
            break

        if code in EMAIL_RETRY_CODES:
            loader.hide()
            QMessageBox.warning(None, "Validação", msg + "\n\nDigite um e-mail válido para continuar.")
            clear_login()
            clear_cache()
            loader.show("Reiniciando validação...")
            continue

        loader.hide()
        QMessageBox.critical(None, "Acesso negado", msg)
        loader.close()
        return

    loader.update("Validando dispositivo (marcador local)...")
    ok_local, reason = ensure_or_mark(allow_create=True)
    if not ok_local:
        loader.hide()
        QMessageBox.critical(None, "Acesso negado", reason)
        loader.close()
        return

    loader.update("Carregando interface...")
    w = App()
    w.show()
    app.processEvents()

    # depende do seu módulo LoadingScreen ter finish_and_close()
    loader.finish_and_close(w)
    app.exec()


if __name__ == "__main__":
    main()