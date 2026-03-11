# excel_io.py
from __future__ import annotations
from datetime import datetime
import os
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from models import Campo, EXCEL_NUMBER_FORMAT
from formatters import cast_value_by_type

def list_sheets_with_ids(path: str) -> List[Tuple[int, str]]:
    """
    Lê um arquivo Excel e retorna uma lista de tuplas contendo o ID da planilha e o nome de cada aba presente no arquivo. O método utiliza a biblioteca openpyxl para carregar o arquivo e iterar sobre as planilhas, extraindo o ID (que pode ser obtido de diferentes atributos dependendo da versão do openpyxl) e o título de cada aba, retornando-os em uma lista estruturada.
    Args:
        path (str): Caminho do arquivo Excel a ser lido.

    Returns:
            list[Tuple[int, str]]: Lista de tuplas, onde cada tupla contém o ID da planilha (int) e o nome da aba (str) presente no arquivo Excel.
    """
    wb = load_workbook(path)
    out: List[Tuple[int, str]] = []
    for ws in wb.worksheets:
        sid = int(getattr(ws, "sheet_id", None) or getattr(ws, "sheetId", None) or ws._id)
        out.append((sid, ws.title))
    return out

def list_sheets(path: str) -> list[str]:
    wb = load_workbook(path)
    return list(wb.sheetnames)

def ensure_sheet_exists(path: str, sheet_name: str) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        wb.save(path)

def is_excel_lock_present(xlsx_path: str) -> bool:
    folder = os.path.dirname(xlsx_path) or "."
    name = os.path.basename(xlsx_path)
    lock_name = "~$" + name
    return os.path.exists(os.path.join(folder, lock_name))


def get_sheet_headers(xlsx_path: str, sheet_name: str) -> list[str]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers: list[str] = []
    for cell in ws[1]:
        v = (cell.value or "")
        v = str(v).strip()
        if v:
            headers.append(v)
    return headers


def read_headers_from_excel(path: str, preferred_sheet: str) -> Tuple[str, List[str], bool]:
    """
    Lê a linha 1 e retorna (sheet_name_usado, headers, has_header).
    has_header = True se encontrou pelo menos 1 valor não vazio na linha 1.
    - Se preferred_sheet não existir, usa a aba ativa.
    """
    wb = load_workbook(path)
    sheet_name = preferred_sheet if preferred_sheet in wb.sheetnames else wb.active.title
    ws = wb[sheet_name]

    if ws.max_row < 1:
        return sheet_name, [], False

    headers: List[str] = []
    has_any = False

    for cell in ws[1]:
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            has_any = True
            if s not in headers:
                headers.append(s)

    return sheet_name, headers, has_any


def write_headers_from_campos(path: str, sheet_name: str, campos: List[Campo]) -> None:
    """
    Garante que a linha 1 tenha cabeçalhos (títulos) baseados em campos do app.
    Se a aba não existir, cria.
    """
    headers = [c.titulo for c in campos]

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    wb.save(path)


def ensure_workbook(path: str, sheet_name: str, headers: List[str]) -> None:
    """
    Garante que o arquivo Excel exista, que a aba exista e que a linha 1 tenha os cabeçalhos (títulos) especificados.
    Args:
        path (str): Caminho do arquivo Excel.
        sheet_name (str): Nome da aba onde os dados serão escritos.
        headers (List[str]): Lista de cabeçalhos (títulos) a serem garantidos na linha 1 da aba. Se a aba já existir, os cabeçalhos serão mesclados com os existentes, mantendo a ordem dos novos e sem duplicatas.
    """
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.max_row < 1:
        ws.append(headers)
        wb.save(path)
        return

    existing = [c.value for c in ws[1]]
    existing = [v for v in existing if v is not None]

    changed = False
    for h in headers:
        if h not in existing:
            existing.append(h)
            changed = True

    if changed:
        for col_idx, h in enumerate(existing, start=1):
            ws.cell(row=1, column=col_idx, value=h)

    wb.save(path)


def apply_column_type_rules(path: str, sheet_name: str, campos: List[Campo]) -> None:
    """
    Aplica regras de formatação e validação de dados nas colunas do Excel com base nos tipos definidos nos campos do app.
    """
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    headers = [h for h in headers if h is not None]
    header_to_col = {h: (headers.index(h) + 1) for h in headers}

    try:
        ws.data_validations.dataValidation = []
    except Exception:
        pass

    max_row = max(ws.max_row, 2)

    for c in campos:
        if c.titulo not in header_to_col:
            continue

        col_idx = header_to_col[c.titulo]
        col_letter = get_column_letter(col_idx)

        if c.tipo == "numero":
            fmt = "General"
        elif c.tipo == "moeda":
            fmt = EXCEL_NUMBER_FORMAT.get("moeda", '"R$" #,##0.00')
        elif c.tipo == "data":
            fmt = EXCEL_NUMBER_FORMAT.get("data", "DD/MM/YYYY")
        else:
            fmt = EXCEL_NUMBER_FORMAT.get(c.tipo, "@")

        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = fmt

        rng = f"{col_letter}2:{col_letter}1048576"

        if c.tipo == "numero":
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = "Digite um número válido (>= 0)."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "moeda":
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = "Digite um valor válido (>= 0)."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "data":
            dv = DataValidation(
                type="date",
                operator="between",
                formula1="DATE(1900,1,1)",
                formula2="DATE(2099,12,31)",
                allow_blank=True,
            )
            dv.errorTitle = "Data inválida"
            dv.error = "Digite uma data válida."
            ws.add_data_validation(dv)
            dv.add(rng)

        elif c.tipo == "booleano":
            dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
            dv.errorTitle = "Valor inválido"
            dv.error = 'Use "Sim" ou "Não".'
            ws.add_data_validation(dv)
            dv.add(rng)

    wb.save(path)
    wb.close()

def _next_data_row(ws) -> int:
    """
    Retorna o índice da próxima linha vazia (base 1) na planilha, considerando a linha 1 como cabeçalho. A função verifica cada linha a partir da linha 2 e considera a linha "ocupada" se qualquer célula nessa linha tiver um valor diferente de None ou vazio. Quando encontra uma linha onde todas as células estão vazias, retorna o índice dessa linha como a próxima linha disponível para inserção de dados.
    Args:
        ws (_type_): Worksheet do openpyxl.

    Returns:
        int: Índice da próxima linha vazia (base 1).
    """
    # começa em 2 (linha 1 é cabeçalho)
    r = 2
    # considera “ocupada” se qualquer célula da linha tiver valor
    while True:
        if any(cell.value not in (None, "") for cell in ws[r]):
            r += 1
            continue
        return r
    
def parse_br_number(value: str):
    """
    Converte texto numérico para int ou float, aceitando padrão brasileiro.

    Exemplos aceitos:
    - 10
    - 10,5
    - 1.234,56

    Retorna:
    - int, se for inteiro
    - float, se tiver parte decimal
    - None, se não for número válido
    """
    value = (value or "").strip()
    if not value:
        return None

    # remove separador de milhar e converte vírgula decimal para ponto
    normalized = value.replace(".", "").replace(",", ".")

    try:
        num = float(normalized)
    except ValueError:
        return None

    if num.is_integer():
        return int(num)

    return num

def parse_br_date_or_text(value: str):
    """
    Retorna datetime se a data for válida e suportada pelo Excel.
    Se for anterior a 1900, retorna o texto original.
    Se não for uma data válida, retorna o texto original.
    """
    value = (value or "").strip()
    if not value:
        return ""

    try:
        dt = datetime.strptime(value, "%d/%m/%Y")
    except ValueError:
        return value

    if dt.year < 1900:
        return value

    return dt

def append_row_typed(file_path: str, sheet_name: str, campos: List[Any], row_by_title: Dict[str, str]) -> None:
    """
    Adiciona uma nova linha na planilha, respeitando o tipo de cada campo.

    Regras principais:
    - texto/email/telefone/... -> grava como texto
    - booleano -> grava como texto
    - numero:
        * grava como int/float
        * sem forçar casas decimais
    - data:
        * se ano >= 1900, grava como datetime com formato DD/MM/YYYY
        * se ano < 1900, grava como texto
    - campos vazios -> célula vazia
    """
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"A aba '{sheet_name}' não existe na planilha.")

    ws = wb[sheet_name]

    if ws.max_row < 1:
        headers = [c.titulo for c in campos]
        ws.append(headers)

    headers_in_sheet = []
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        headers_in_sheet.append("" if cell_value is None else str(cell_value).strip())

    header_to_col = {title: idx + 1 for idx, title in enumerate(headers_in_sheet) if title}

    for campo in campos:
        if campo.titulo not in header_to_col:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = campo.titulo
            header_to_col[campo.titulo] = new_col

    next_row = ws.max_row + 1

    for campo in campos:
        col_idx = header_to_col[campo.titulo]
        cell = ws.cell(row=next_row, column=col_idx)

        raw_value = row_by_title.get(campo.titulo, "")
        raw_value = "" if raw_value is None else str(raw_value).strip()

        if raw_value == "":
            cell.value = ""
            cell.number_format = "General"
            continue

        if campo.tipo == "data":
            parsed = parse_br_date_or_text(raw_value)

            if isinstance(parsed, datetime):
                cell.value = parsed
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = parsed
                cell.number_format = "@"

        elif campo.tipo == "numero":
            parsed_num = parse_br_number(raw_value)

            if parsed_num is None:
                # fallback defensivo: grava como texto se vier algo inválido
                cell.value = raw_value
                cell.number_format = "@"
            else:
                cell.value = parsed_num
                cell.number_format = "General"

        elif campo.tipo == "moeda":
            parsed_num = parse_br_number(raw_value)

            if parsed_num is None:
                cell.value = raw_value
                cell.number_format = "@"
            else:
                cell.value = float(parsed_num)
                cell.number_format = EXCEL_NUMBER_FORMAT.get("moeda", '"R$" #,##0.00')

        elif campo.tipo == "booleano":
            cell.value = raw_value
            cell.number_format = "@"

        else:
            cell.value = raw_value
            cell.number_format = "@"

    wb.save(file_path)
    wb.close()

def delete_column_by_header(path: str, sheet_name: str, header_name: str) -> None:
    """
    Exclui uma coluna do arquivo Excel com base no nome do cabeçalho. O método verifica se o arquivo e a aba existem, se a linha 1 contém o cabeçalho especificado e, se todas as condições forem atendidas, exclui a coluna correspondente ao cabeçalho fornecido.
    Args:
        path (str): Caminho do arquivo Excel.
        sheet_name (str): Nome da aba onde a coluna será excluída.
        header_name (str): Nome do cabeçalho da coluna a ser excluída.
    """
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    if header_name not in headers:
        wb.save(path)
        return

    col = headers.index(header_name) + 1
    ws.delete_cols(col, 1)
    wb.save(path)


def rename_column_header(path: str, sheet_name: str, old_header: str, new_header: str) -> None:
    """
    Renomeia o cabeçalho de uma coluna no arquivo Excel. O método verifica se o arquivo e a aba existem, se a linha 1 contém o cabeçalho antigo especificado e, se todas as condições forem atendidas, atualiza o valor do cabeçalho para o novo nome fornecido.
    Args:
        path (str): Caminho do arquivo Excel.
        sheet_name (str): Nome da aba onde a coluna será renomeada.
        old_header (str): Nome do cabeçalho atual da coluna a ser renomeada.
        new_header (str): Novo nome do cabeçalho da coluna.
    """
    if not path or not os.path.exists(path):
        return

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    if ws.max_row < 1:
        wb.save(path)
        return

    headers = [c.value for c in ws[1]]
    if old_header not in headers:
        wb.save(path)
        return

    col = headers.index(old_header) + 1
    ws.cell(row=1, column=col, value=new_header)
    wb.save(path)